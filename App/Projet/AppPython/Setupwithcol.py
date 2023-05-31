import pandas as pd
import mysql.connector as mysql
import openpyxl
from numpy import delete, argwhere
from time import sleep
from tqdm import tqdm


def fillDB(self, path_to_excel, _type):
    # La warwait
    
    self.HOST="10.0.1.238"
    self.DATABASE=["cis_nord_warwait","skillmatrix"]
    self.USER="admin"
    self.PASSWORD="clecle0202."
    
    if _type == "ww":
        if self.all == 0:
            try:
                numweek = int(self.weeknumEntry.get())
            except ValueError as ve:
                raise ValueError("Enter an integer for the week's num please.")
            except Exception as e:
                raise Exception("Error when converting the week's num, please try again.")

        try:
            database = mysql.connect(
                host=self.HOST,
                database=self.DATABASE[0],
                user=self.USER,
                password=self.PASSWORD
            )
            print("Connected to database... OK")
        except mysql.Error as e:
            raise Exception("Can't connect to the database: {}".format(e))

        cursor = database.cursor()
        print("test")
        if self.all == 0:
            try:
                week = pd.read_excel(path_to_excel, sheet_name=numweek, skiprows=8, nrows=1)
                db = pd.read_excel(path_to_excel, sheet_name=numweek, skiprows=9)

                xl = pd.ExcelFile(path_to_excel).sheet_names[numweek]
                wb = openpyxl.load_workbook(path_to_excel)
                fs = wb[xl]
                fs_count_row = fs.max_row
                fs_count_col = fs.max_column

                for row in range(1, fs_count_row + 1):
                    for column in range(1, fs_count_col + 1):
                        cell_color = fs.cell(column=column, row=row)
                        bgColor = cell_color.fill.bgColor.index
                        fgColor = cell_color.fill.fgColor.index

                        if (bgColor == '00000000') or (fgColor == '00000000'):
                            print("true")
                            continue
                        else:
                            print("Background color index of cell (", row, column, ") is", bgColor)
                            print("Foreground color index of cell (", row, column, ") is", fgColor)

            except PermissionError as e:
                raise PermissionError("Can't open the file, it is already running.")

            lis = db.to_numpy()
            week_list = week.to_numpy()

            create_table_warwait = """CREATE TABLE IF NOT EXISTS warwait (nom TEXT, grade TEXT, site TEXT,s1 TEXT,s2 TEXT,s3 TEXT,s4 TEXT,s5 TEXT,s6 TEXT,s7 TEXT,s8 TEXT,s9 TEXT,s10 TEXT,s11 TEXT,s12 TEXT,s13 TEXT,s14 TEXT,s15 TEXT,s16 TEXT,s17 TEXT,s18 TEXT,s19 TEXT,s20 TEXT,s21 TEXT,s22 TEXT,s23 TEXT,s24 TEXT,s25 TEXT,s26 TEXT,s27 TEXT,s28 TEXT,s29 TEXT,s30 TEXT,s31 TEXT,s32 TEXT,s33 TEXT,s34 TEXT,s35 TEXT,s36 TEXT,s37 TEXT,s38 TEXT,s39 TEXT,s40 TEXT,s41 TEXT,s42 TEXT,s43 TEXT,s44 TEXT,s45 TEXT,s46 TEXT,s47 TEXT,s48 TEXT,s49 TEXT,s50 TEXT,s51 TEXT,s52 TEXT,reussite FLOAT,positionnement TEXT,competences TEXT,cv_code INTEGER,pe TEXT,en_mission TEXT,afficher TEXT,id integer auto_increment primary key)"""
            try:
                cursor.execute(create_table_warwait)
            except mysql.Error as e:
                raise Exception("Can't create the table: {}".format(e))

            database.commit()

            # taking weeks number on warwait and throwing away all empty cells in the.
            # prenant les numéros de semaine sur warwait et en jetant toutes les cellules vides dans le chemin numweek
            week_list = week_list[~pd.isnull(week_list)]
            week_list = week_list.tolist()
            week_list = list(map(int, week_list))

            week_range = range(len(week_list))

            # remove une semaine si deja dans la db (a verifier)
            for x in tqdm(week_range, desc="Processing weeks"):
                val = week_list[x]

                check_week = "SELECT * FROM warwait WHERE s1 = {}".format(val)

                try:
                    cursor.execute(check_week)
                    if cursor.fetchone():
                        week_list.remove(val)
                except mysql.Error as e:
                    raise Exception("Error checking week in database: {}".format(e))

            # supprimer les colonnes nulles (a verifier)
            delete(db, argwhere(pd.isnull(db)))
            db = pd.DataFrame(db)

            db.columns = db.iloc[0]
            db = db.drop(0)

            if week_list:
                for i in tqdm(range(len(db)), desc="Filling database"):
                    val = db.iloc[i].values.tolist()
                    val = list(map(str, val))
                    val = [None if x == 'nan' else x for x in val]

                    try:
                        cursor.execute(
                            "INSERT INTO warwait(nom, grade, site, s1, s2, s3, s4, s5, s6, s7, s8, s9, s10, s11, s12, s13, s14, s15, s16, s17, s18, s19, s20, s21, s22, s23, s24, s25, s26, s27, s28, s29, s30, s31, s32, s33, s34, s35, s36, s37, s38, s39, s40, s41, s42, s43, s44, s45, s46, s47, s48, s49, s50, s51, s52, reussite, positionnement, competences, cv_code, pe, en_mission, afficher) VALUES (%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s,%s)",
                            tuple(val))
                    except mysql.Error as e:
                        raise Exception("Error inserting values into database: {}".format(e))

                print("All weeks have been processed.")
            else:
                print("All weeks are already in the database.")

            database.commit()
            cursor.close()
            database.close()

    # Fin warwait

# Exemple d'utilisation
# path_to_excel = "../Wait_Room_2023.xlsx"
# _type = "ww"
# fillDB(path_to_excel, _type)
